import os
from typing import Dict, Any
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
import logging

logger = logging.getLogger(__name__)

class DocumentGenerator:
    def __init__(self, output_dir: str = None):
        if output_dir is None:
            # Utiliser le chemin absolu par défaut
            project_root = Path(__file__).parent.parent.parent.resolve()
            self.output_dir = os.path.join(project_root, "generated_business_plans")
        else:
            self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
    
    def generate_excel_business_plan(self, business_plan_data: Dict[str, Any], filename: str = None) -> str:
        """Génère un fichier Excel complet du business plan."""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"business_plan_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Supprimer la feuille par défaut
            
            # Styles communs
            header_font = Font(size=14, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            subheader_font = Font(size=12, bold=True)
            normal_font = Font(size=10)
            
            # 1. Feuille Résumé Exécutif
            ws_resume = wb.create_sheet("Résumé Exécutif")
            self._create_excel_resume_sheet(ws_resume, business_plan_data, header_font, header_fill, subheader_font)
            
            # 2. Feuille Analyse Marché
            ws_marche = wb.create_sheet("Analyse Marché")
            self._create_excel_marche_sheet(ws_marche, business_plan_data, header_font, header_fill, subheader_font)
            
            # 3. Feuille Projections Financières
            ws_finance = wb.create_sheet("Projections Financières")
            self._create_excel_finance_sheet(ws_finance, business_plan_data, header_font, header_fill, subheader_font)
            
            # 4. Feuille Stratégie Marketing
            ws_marketing = wb.create_sheet("Stratégie Marketing")
            self._create_excel_marketing_sheet(ws_marketing, business_plan_data, header_font, header_fill, subheader_font)
            
            # 5. Feuille Plan Opérationnel
            ws_ops = wb.create_sheet("Plan Opérationnel")
            self._create_excel_operations_sheet(ws_ops, business_plan_data, header_font, header_fill, subheader_font)
            
            # 6. Feuille Risques & Opportunités
            ws_risks = wb.create_sheet("Risques & Opportunités")
            self._create_excel_risks_sheet(ws_risks, business_plan_data, header_font, header_fill, subheader_font)
            
            wb.save(filepath)
            logger.info(f"Fichier Excel généré: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Erreur génération Excel: {str(e)}")
            raise
    
    def generate_pdf_business_plan(self, business_plan_data: Dict[str, Any], filename: str = None) -> str:
        """Génère un fichier PDF de l'itinéraire technique."""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"itineraire_technique_{timestamp}.pdf"
        
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Styles personnalisés
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontSize=24,
                spaceAfter=30,
                textColor=colors.HexColor('#366092')
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=12,
                textColor=colors.HexColor('#366092')
            )
            
            # Page de titre
            story.append(Paragraph(f"ITINÉRAIRE TECHNIQUE - {business_plan_data.get('titre', 'Projet')}", title_style))
            story.append(Spacer(1, 0.5*inch))
            story.append(Paragraph(f"Document technique généré le {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
            story.append(Spacer(1, 1*inch))
            
            # Table des matières
            story.append(Paragraph("Table des Matières", heading_style))
            toc_data = [
                ["1. Architecture Technique", "3"],
                ["2. Spécifications Détaillées", "4"],
                ["3. Étapes de Développement", "5"],
                ["4. Planning d'Implémentation", "6"],
                ["5. Ressources Techniques", "7"],
                ["6. Technologies Recommandées", "8"],
                ["7. Contraintes et Solutions", "9"],
                ["8. Recommandations Techniques", "10"]
            ]
            
            toc_table = Table(toc_data, colWidths=[4*inch, 1*inch])
            toc_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
                ('FONTSIZE', (0,0), (-1,-1), 10),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ]))
            story.append(toc_table)
            story.append(Spacer(1, 1*inch))
            
            # Contenu technique détaillé
            itineraire = business_plan_data.get('itineraire_technique', {})
            self._add_pdf_technical_section(story, "1. ARCHITECTURE TECHNIQUE", itineraire.get('architecture', ''), styles, heading_style)
            self._add_pdf_technical_section(story, "2. SPÉCIFICATIONS DÉTAILLÉES", itineraire.get('specifications', ''), styles, heading_style)
            self._add_pdf_technical_section(story, "3. ÉTAPES DE DÉVELOPPEMENT", itineraire.get('etapes_developpement', ''), styles, heading_style)
            self._add_pdf_technical_section(story, "4. PLANNING D'IMPLÉMENTATION", itineraire.get('planning_implementation', ''), styles, heading_style)
            self._add_pdf_technical_section(story, "5. RESSOURCES TECHNIQUES", itineraire.get('ressources_techniques', ''), styles, heading_style)
            self._add_pdf_technical_section(story, "6. TECHNOLOGIES RECOMMANDÉES", itineraire.get('technologies', ''), styles, heading_style)
            self._add_pdf_technical_section(story, "7. CONTRAINTES ET SOLUTIONS", itineraire.get('contraintes', ''), styles, heading_style)
            self._add_pdf_section(story, "8. RECOMMANDATIONS TECHNIQUES", business_plan_data.get('recommandations', {}), styles, heading_style)
            
            doc.build(story)
            logger.info(f"Fichier PDF généré: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Erreur génération PDF: {str(e)}")
            raise
    
    def _create_excel_resume_sheet(self, ws, data, header_font, header_fill, subheader_font):
        """Crée la feuille résumé exécutif Excel."""
        ws['A1'] = "RÉSUMÉ EXÉCUTIF"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:E1')
        
        row = 3
        resume_data = data.get('resume_executif', {})
        for key, value in resume_data.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'A{row}'].font = subheader_font
            ws[f'B{row}'] = str(value)
            ws.merge_cells(f'B{row}:E{row}')
            row += 2
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def _create_excel_finance_sheet(self, ws, data, header_font, header_fill, subheader_font):
        """Crée la feuille projections financières Excel."""
        ws['A1'] = "PROJECTIONS FINANCIÈRES"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:E1')
        
        finance_data = data.get('projections_financieres', {})
        
        # Tableau compte de résultat
        if 'compte_resultat_3ans' in finance_data:
            ws['A3'] = "COMPTE DE RÉSULTAT PRÉVISIONNEL (3 ans)"
            ws['A3'].font = subheader_font
            
            headers = ['Éléments', 'Année 1 (€)', 'Année 2 (€)', 'Année 3 (€)']
            for i, header in enumerate(headers, 1):
                ws.cell(row=5, column=i, value=header)
                ws.cell(row=5, column=i).font = subheader_font
                ws.cell(row=5, column=i).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            
            compte_resultat = finance_data['compte_resultat_3ans']
            rows_data = [
                ['Chiffre d\'affaires', 
                 compte_resultat.get('annee_1', {}).get('chiffre_affaires', 0),
                 compte_resultat.get('annee_2', {}).get('chiffre_affaires', 0),
                 compte_resultat.get('annee_3', {}).get('chiffre_affaires', 0)],
                ['Charges totales',
                 compte_resultat.get('annee_1', {}).get('charges', 0),
                 compte_resultat.get('annee_2', {}).get('charges', 0),
                 compte_resultat.get('annee_3', {}).get('charges', 0)],
                ['Résultat net',
                 compte_resultat.get('annee_1', {}).get('resultat', 0),
                 compte_resultat.get('annee_2', {}).get('resultat', 0),
                 compte_resultat.get('annee_3', {}).get('resultat', 0)]
            ]
            
            for i, row_data in enumerate(rows_data, 6):
                for j, value in enumerate(row_data, 1):
                    cell = ws.cell(row=i, column=j, value=value)
                    if j > 1 and isinstance(value, (int, float)):
                        cell.number_format = '#,##0'
        
        # Plan de financement
        if 'plan_financement' in finance_data:
            ws['A12'] = "PLAN DE FINANCEMENT"
            ws['A12'].font = subheader_font
            
            row = 14
            for key, value in finance_data['plan_financement'].items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = str(value)
                row += 1
    
    def _add_pdf_section(self, story, title, data, styles, heading_style):
        """Ajoute une section au PDF."""
        story.append(Paragraph(title, heading_style))
        story.append(Spacer(1, 12))
        
        for key, value in data.items():
            subtitle = key.replace('_', ' ').title()
            story.append(Paragraph(f"<b>{subtitle}:</b>", styles['Normal']))
            story.append(Paragraph(str(value), styles['Normal']))
            story.append(Spacer(1, 6))
        
        story.append(Spacer(1, 24))
    
    def _add_pdf_technical_section(self, story, title, content, styles, heading_style):
        """Ajoute une section technique au PDF."""
        story.append(Paragraph(title, heading_style))
        story.append(Spacer(1, 12))
        
        if content:
            # Traiter le contenu technique avec formatage spécial
            paragraphs = content.split('\n')
            for paragraph in paragraphs:
                if paragraph.strip():
                    # Détecter les listes à puces
                    if paragraph.strip().startswith('-') or paragraph.strip().startswith('•'):
                        story.append(Paragraph(f"    {paragraph.strip()}", styles['Normal']))
                    # Détecter les titres (commence par majuscule et se termine par :)
                    elif ':' in paragraph and paragraph.strip().endswith(':'):
                        story.append(Paragraph(f"<b>{paragraph.strip()}</b>", styles['Normal']))
                    else:
                        story.append(Paragraph(paragraph.strip(), styles['Normal']))
                    story.append(Spacer(1, 6))
        else:
            story.append(Paragraph("Section à compléter selon les spécificités du projet.", styles['Normal']))
            story.append(Spacer(1, 6))
        
        story.append(Spacer(1, 24))
    
    def _add_pdf_finance_section(self, story, title, data, styles, heading_style):
        """Ajoute la section financière spécialisée au PDF."""
        story.append(Paragraph(title, heading_style))
        story.append(Spacer(1, 12))
        
        # Tableau compte de résultat
        if 'compte_resultat_3ans' in data:
            story.append(Paragraph("<b>Compte de Résultat Prévisionnel (3 ans)</b>", styles['Normal']))
            story.append(Spacer(1, 6))
            
            compte_data = data['compte_resultat_3ans']
            table_data = [
                ['Éléments', 'Année 1 (€)', 'Année 2 (€)', 'Année 3 (€)'],
                ['Chiffre d\'affaires',
                 f"{compte_data.get('annee_1', {}).get('chiffre_affaires', 0):,}",
                 f"{compte_data.get('annee_2', {}).get('chiffre_affaires', 0):,}",
                 f"{compte_data.get('annee_3', {}).get('chiffre_affaires', 0):,}"],
                ['Charges totales',
                 f"{compte_data.get('annee_1', {}).get('charges', 0):,}",
                 f"{compte_data.get('annee_2', {}).get('charges', 0):,}",
                 f"{compte_data.get('annee_3', {}).get('charges', 0):,}"],
                ['Résultat net',
                 f"{compte_data.get('annee_1', {}).get('resultat', 0):,}",
                 f"{compte_data.get('annee_2', {}).get('resultat', 0):,}",
                 f"{compte_data.get('annee_3', {}).get('resultat', 0):,}"]
            ]
            
            table = Table(table_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.grey),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 10),
                ('BOTTOMPADDING', (0,0), (-1,0), 12),
                ('BACKGROUND', (0,1), (-1,-1), colors.beige),
                ('GRID', (0,0), (-1,-1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 12))
        
        # Autres éléments financiers
        for key, value in data.items():
            if key != 'compte_resultat_3ans':
                subtitle = key.replace('_', ' ').title()
                story.append(Paragraph(f"<b>{subtitle}:</b>", styles['Normal']))
                story.append(Paragraph(str(value), styles['Normal']))
                story.append(Spacer(1, 6))
        
        story.append(Spacer(1, 24))
    
    def _create_excel_marche_sheet(self, ws, data, header_font, header_fill, subheader_font):
        """Crée la feuille analyse marché Excel."""
        ws['A1'] = "ANALYSE DU MARCHÉ"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:E1')
        
        row = 3
        marche_data = data.get('analyse_marche', {})
        for key, value in marche_data.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'A{row}'].font = subheader_font
            ws[f'B{row}'] = str(value)
            ws.merge_cells(f'B{row}:E{row}')
            row += 2
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def _create_excel_marketing_sheet(self, ws, data, header_font, header_fill, subheader_font):
        """Crée la feuille stratégie marketing Excel."""
        ws['A1'] = "STRATÉGIE MARKETING"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:E1')
        
        row = 3
        marketing_data = data.get('strategie_marketing', {})
        for key, value in marketing_data.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'A{row}'].font = subheader_font
            ws[f'B{row}'] = str(value)
            ws.merge_cells(f'B{row}:E{row}')
            row += 2
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def _create_excel_operations_sheet(self, ws, data, header_font, header_fill, subheader_font):
        """Crée la feuille plan opérationnel Excel."""
        ws['A1'] = "PLAN OPÉRATIONNEL"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:E1')
        
        row = 3
        ops_data = data.get('plan_operationnel', {})
        for key, value in ops_data.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'A{row}'].font = subheader_font
            ws[f'B{row}'] = str(value)
            ws.merge_cells(f'B{row}:E{row}')
            row += 2
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def _create_excel_risks_sheet(self, ws, data, header_font, header_fill, subheader_font):
        """Crée la feuille risques et opportunités Excel."""
        ws['A1'] = "RISQUES & OPPORTUNITÉS"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:E1')
        
        row = 3
        risks_data = data.get('risques_opportunites', {})
        for key, value in risks_data.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'A{row}'].font = subheader_font
            ws[f'B{row}'] = str(value)
            ws.merge_cells(f'B{row}:E{row}')
            row += 2
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
